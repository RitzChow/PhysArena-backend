from __future__ import annotations

import json
import math
import os
import glob
from collections import defaultdict, OrderedDict
from dataclasses import dataclass
from typing import Dict, List, Any, Tuple
from flask_cors import CORS
from flask import Flask, jsonify, send_from_directory
from openpyxl import load_workbook


# -------------------------
# Configuration
# -------------------------
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

# 竞赛配置 - 支持多个竞赛文件
COMPETITION_CONFIGS = {
    "ipho--ipho_2024": {
        "nice_name": "IPhO 2024",
        "type": "FinalAnswer",
        "medal_thresholds": [75, 50, 25],
        "judge": False,
        "default_open": True,
        "files": ["ipho2024_outputs.xlsx"]
    },
    "ipho--ipho_2025": {
        "nice_name": "IPhO 2025", 
        "type": "FinalAnswer",
        "medal_thresholds": [75, 50, 25],
        "judge": False,
        "default_open": False,
        "files": ["ipho2025_outputs.xlsx"]
    }
    # 可以继续添加更多竞赛配置
}

# 自动发现所有xlsx文件
def discover_competition_files():
    """自动发现项目目录中的所有竞赛输出文件"""
    xlsx_files = glob.glob(os.path.join(PROJECT_ROOT, "*_outputs.xlsx"))
    discovered_configs = {}
    
    for file_path in xlsx_files:
        filename = os.path.basename(file_path)
        # 从文件名推断竞赛名称
        if "ipho2024" in filename:
            key = "ipho--ipho_2024"
            nice_name = "IPhO 2024"
        elif "ipho2025" in filename:
            key = "ipho--ipho_2025"
            nice_name = "IPhO 2025"
        else:
            # 通用处理，从文件名推断
            base_name = filename.replace("_outputs.xlsx", "")
            key = f"{base_name}--{base_name}"
            nice_name = base_name.upper()
        
        if key not in discovered_configs:
            discovered_configs[key] = {
                "nice_name": nice_name,
                "type": "FinalAnswer",
                "medal_thresholds": [75, 50, 25],
                "judge": False,
                "default_open": len(discovered_configs) == 0,  # 第一个发现的设为默认
                "files": [filename]
            }
        else:
            discovered_configs[key]["files"].append(filename)
    
    return discovered_configs

# 合并配置
COMPETITION_CONFIGS.update(discover_competition_files())


@dataclass
class RunRecord:
    problem_idx: str
    problem_statement: str
    model_name: str
    model_config: str
    idx_answer: int
    user_message: str
    answer: str
    messages: str
    input_tokens: float
    output_tokens: float
    run_cost: float
    input_cost_per_tokens: float
    output_cost_per_tokens: float
    gold_answer: str | None
    parsed_answer: str | None
    correct: bool | None


def read_outputs_xlsx(path: str) -> List[RunRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    def get(row, key, default=None):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else default

    data: List[RunRecord] = []
    for row in rows[1:]:
        if not any(x is not None for x in row):
            continue
        try:
            record = RunRecord(
                problem_idx=str(get(row, "problem_idx")),
                problem_statement=str(get(row, "problem")),
                model_name=str(get(row, "model_name")),
                model_config=str(get(row, "model_config")),
                idx_answer=int(get(row, "idx_answer", 0) or 0),
                user_message=str(get(row, "user_message")),
                answer=str(get(row, "answer")),
                messages=str(get(row, "messages")),
                input_tokens=float(get(row, "input_tokens", 0) or 0),
                output_tokens=float(get(row, "output_tokens", 0) or 0),
                run_cost=float(get(row, "cost", 0) or 0),
                input_cost_per_tokens=float(get(row, "input_cost_per_tokens", 0) or 0),
                output_cost_per_tokens=float(get(row, "output_cost_per_tokens", 0) or 0),
                gold_answer=(get(row, "gold_answer") if get(row, "gold_answer") is not None else None),
                parsed_answer=(get(row, "parsed_answer") if get(row, "parsed_answer") is not None else None),
                correct=bool(get(row, "correct")) if get(row, "correct") is not None else None,
            )
            data.append(record)
        except Exception:
            # Skip malformed row but continue
            continue
    return data


def build_backend_payload() -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
    all_results_payload = {}
    all_secondary_payload = {}
    all_competition_dates = {}
    all_traces_index = {}
    
    for competition_key, config in COMPETITION_CONFIGS.items():
        # 读取该竞赛的所有文件
        all_runs = []
        for filename in config["files"]:
            file_path = os.path.join(PROJECT_ROOT, filename)
            if os.path.exists(file_path):
                runs = read_outputs_xlsx(file_path)
                all_runs.extend(runs)
        
        if not all_runs:
            continue
            
        # 从运行记录中提取问题ID和顺序
        problem_ids_in_order = []
        problem_id_to_gold = OrderedDict()
        
        # 按模型和问题ID分组
        grouped: Dict[Tuple[str, str], List[RunRecord]] = defaultdict(list)
        model_set: set[str] = set()
        
        for r in all_runs:
            model_set.add(r.model_name)
            grouped[(r.model_name, r.problem_idx)].append(r)
            
            # 收集问题ID和标准答案
            if r.problem_idx not in problem_ids_in_order:
                problem_ids_in_order.append(r.problem_idx)
                if r.gold_answer is not None:
                    problem_id_to_gold[r.problem_idx] = r.gold_answer
                else:
                    problem_id_to_gold[r.problem_idx] = ""

        # 聚合每个模型的总计
        model_totals = {m: {"input_tokens": 0.0, "output_tokens": 0.0, "cost": 0.0} for m in model_set}
        model_price: Dict[str, Dict[str, float]] = {m: {"input": None, "output": None} for m in model_set}

        for r in all_runs:
            mt = model_totals[r.model_name]
            mt["input_tokens"] += r.input_tokens or 0
            mt["output_tokens"] += r.output_tokens or 0
            mt["cost"] += r.run_cost or 0
            # 保存首次看到的价格
            if model_price[r.model_name]["input"] is None and r.input_cost_per_tokens is not None:
                v = float(r.input_cost_per_tokens or 0)
                model_price[r.model_name]["input"] = v
            if model_price[r.model_name]["output"] is None and r.output_cost_per_tokens is not None:
                v = float(r.output_cost_per_tokens or 0)
                model_price[r.model_name]["output"] = v

        # 构建结果行
        numeric_to_pid: Dict[int, str] = {}
        problem_names: List[str] = []
        for i, pid in enumerate(problem_ids_in_order, start=1):
            numeric_to_pid[i] = pid
            problem_names.append(pid)

        # 为每个数值任务计算每个模型的准确率
        results_rows: List[Dict[str, Any]] = []
        for idx in range(1, len(problem_ids_in_order) + 1):
            pid = numeric_to_pid[idx]
            row: Dict[str, Any] = {"question": idx}
            for m in sorted(model_set):
                runs_for = grouped.get((m, pid), [])
                if not runs_for:
                    row[m] = 0
                else:
                    num = sum(1 for r in runs_for if (r.correct is True))
                    den = len(runs_for)
                    acc = 100.0 * num / den if den > 0 else 0.0
                    row[m] = acc
            results_rows.append(row)

        # 平均行
        avg_row: Dict[str, Any] = {"question": "Avg"}
        for m in sorted(model_set):
            vals = [r[m] for r in results_rows if isinstance(r[m], (int, float))]
            avg_row[m] = sum(vals) / len(vals) if vals else 0.0
        results_rows.append(avg_row)

        # 成本行
        cost_row: Dict[str, Any] = {"question": "Cost"}
        for m in sorted(model_set):
            cost_row[m] = model_totals[m]["cost"]
        results_rows.append(cost_row)

        # 构建该竞赛的结果负载
        all_results_payload[competition_key] = results_rows

        # 构建次要负载行
        secondary_rows: List[Dict[str, Any]] = []
        # 输入令牌
        row_it = {"question": "Input Tokens"}
        for m in sorted(model_set):
            row_it[m] = model_totals[m]["input_tokens"]
        secondary_rows.append(row_it)

        # 输入成本
        row_icpt = {"question": "Input Cost"}
        for m in sorted(model_set):
            price_per_mtok = model_price[m]["input"] or 0.0
            tokens = model_totals[m]["input_tokens"] or 0.0
            dollars = (tokens * price_per_mtok) / 1_000_000.0
            row_icpt[m] = round(dollars, 6)
        secondary_rows.append(row_icpt)

        # 输出令牌
        row_ot = {"question": "Output Tokens"}
        for m in sorted(model_set):
            row_ot[m] = model_totals[m]["output_tokens"]
        secondary_rows.append(row_ot)

        # 输出成本
        row_ocpt = {"question": "Output Cost"}
        for m in sorted(model_set):
            price_per_mtok = model_price[m]["output"] or 0.0
            tokens = model_totals[m]["output_tokens"] or 0.0
            dollars = (tokens * price_per_mtok) / 1_000_000.0
            row_ocpt[m] = round(dollars, 6)
        secondary_rows.append(row_ocpt)

        # 准确率
        row_acc = {"question": "Acc"}
        for m in sorted(model_set):
            row_acc[m] = avg_row[m]
        secondary_rows.append(row_acc)

        all_secondary_payload[competition_key] = secondary_rows

        # 构建竞赛日期：全部设为False（无污染警告）
        all_competition_dates[competition_key] = {}
        for m in model_set:
            all_competition_dates[competition_key][m] = False

        # 构建跟踪索引
        traces_index: Dict[Tuple[str, int], Dict[str, Any]] = {}
        for idx in range(1, len(problem_ids_in_order) + 1):
            pid = numeric_to_pid[idx]
            gold = problem_id_to_gold.get(pid, "")
            for m in model_set:
                runs_for = sorted(grouped.get((m, pid), []), key=lambda r: r.idx_answer)
                if not runs_for:
                    continue
                statement = runs_for[0].problem_statement or ""
                outs: List[Dict[str, Any]] = []
                for rr in runs_for:
                    outs.append({
                        "parsed_answer": rr.parsed_answer,
                        "correct": bool(rr.correct) if rr.correct is not None else False,
                        "solution": rr.answer or "",
                    })
                traces_index[(m, idx)] = {
                    "statement": statement,
                    "gold_answer": gold,
                    "model_outputs": outs,
                }
        
        all_traces_index[competition_key] = traces_index

    # 构建顶级结果负载
    competition_info = {}
    for i, (competition_key, config) in enumerate(COMPETITION_CONFIGS.items(), start=1):
        if competition_key in all_results_payload:
            competition_info[competition_key] = {
                "index": i,
                "nice_name": config["nice_name"],
                "type": config["type"],
                "num_problems": len(all_results_payload[competition_key]) - 2,  # 减去Avg和Cost行
                "medal_thresholds": config["medal_thresholds"],
                "judge": config["judge"],
                "default_open": config.get("default_open", False),
                "problem_names": [f"{i+1}" for i in range(len(all_results_payload[competition_key]) - 2)]
            }

    results_payload = {
        "competition_info": competition_info,
        "results": all_results_payload
    }

    return results_payload, all_secondary_payload, all_competition_dates, all_traces_index


# -------------------------
# Flask app
# -------------------------
app = Flask(__name__, static_folder=PROJECT_ROOT, static_url_path="")
CORS(app)
RESULTS_PAYLOAD, SECONDARY_PAYLOAD, COMP_DATES_PAYLOAD, TRACES_INDEX = build_backend_payload()


@app.route("/")
def root():
    return send_from_directory(PROJECT_ROOT, "index.html")


@app.get("/results")
def get_results():
    return jsonify(RESULTS_PAYLOAD)


@app.get("/secondary")
def get_secondary():
    return jsonify(SECONDARY_PAYLOAD)


@app.get("/competition_dates")
def get_comp_dates():
    return jsonify(COMP_DATES_PAYLOAD)


@app.get("/traces/<competition>/<model>/<int:task>")
def get_traces(competition: str, model: str, task: int):
    if competition not in TRACES_INDEX:
        return jsonify({"error": "competition not found"}), 404
    key = (model, task)
    data = TRACES_INDEX[competition].get(key)
    if not data:
        return jsonify({"error": "trace not found"}), 404
    return jsonify(data)


def main():
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=True)


if __name__ == "__main__":
    main()


